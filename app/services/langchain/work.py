import os
import asyncio
import json
import re
import unicodedata
from typing import Optional
from openai import OpenAI
from openpyxl import load_workbook
from pathlib import Path   
from app.services.langchain.prompts import *
from app.utils.json_formatter import clean_and_parse_json
from app.core.config import settings
import csv

client = OpenAI(api_key=settings.OPENAI_API_KEY)

ASSISTANT_ID = "asst_dWlfRIYPk4VmNDGOB9rY78iF"

BASE_DIR = Path(__file__).resolve().parent  # app/services/langchain
EXCEL_PATH = BASE_DIR / "data" / "meteriality-map.xlsx"


MIN_ROWS_PROMPT_2 = 10
MAX_ROWS_PROMPT_2 = 30


# ==================================================
# JSON helpers
# ==================================================
def try_fix_json(raw_text: str):
    if not raw_text:
        return None

    # Normalizar comillas raras
    raw_text = raw_text.replace("“", '"').replace("”", '"').replace("’", "'")

    # 1) Intentar extraer solo lo que está ENTRE <JSON>...</JSON>
    m = re.search(r"<JSON>(.*)</JSON>", raw_text, re.DOTALL | re.IGNORECASE)
    if m:
        raw_text = m.group(1).strip()
    else:
        # 2) Fallback viejo: buscar el primer bloque {...}
        json_candidate = re.search(r"\{.*\}", raw_text, re.DOTALL)
        if json_candidate:
            raw_text = json_candidate.group(0)

    # Limpieza básica
    raw_text = raw_text.replace("\n", " ").replace("\t", " ")
    raw_text = re.sub(r",(\s*[}\]])", r"\1", raw_text)

    # 3) Intento directo
    try:
        return json.loads(raw_text)
    except Exception:
        pass

    # 4) Fallback avanzado (tu clean_and_parse_json)
    try:
        return clean_and_parse_json(raw_text)
    except Exception:
        return None

def try_fix_json_prompt6(raw_text: str):
    """
    Parser especial para Prompt 6:
    - Intenta primero extraer lo que esté entre <JSON>...</JSON>
    - Si no hay etiquetas, cae al parser genérico (try_fix_json)
    """
    if not raw_text:
        return None

    m = re.search(r"<JSON>(.*)</JSON>", raw_text, re.DOTALL | re.IGNORECASE)
    if m:
        payload = m.group(1).strip()
    else:
        return try_fix_json(raw_text)

    try:
        return json.loads(payload)
    except Exception:
        pass

    try:
        return clean_and_parse_json(payload)
    except Exception:
        return None


def try_fix_json_prompt10(raw_text: str):
    """
    Parser específico para Prompt 10.
    - Soporta JSON puro.
    - Soporta JSON envuelto en ```json ... ```.
    - Siempre intenta extraer desde la primera '{' hasta la última '}'.
    """
    if not raw_text:
        return None

    cleaned = raw_text.strip()

    # 1) Si viene con ```json ... ```
    if cleaned.startswith("```"):
        m = re.search(r"```(?:json)?(.*)```", cleaned, re.DOTALL | re.IGNORECASE)
        if m:
            cleaned = m.group(1).strip()

    # 2) Extraer solo el bloque {...}
    m2 = re.search(r"\{.*\}", cleaned, re.DOTALL)
    if m2:
        cleaned = m2.group(0).strip()

    # 3) Intento directo
    try:
        return json.loads(cleaned)
    except Exception as e:
        print("❌ Error parseando JSON de Prompt 10:", e)
        print("📄 Fragmento problemático (primeros 800 chars):")
        print(repr(cleaned[:800]))
        return None


def normalize_json_keys(obj):
    if isinstance(obj, dict):
        new = {}
        for k, v in obj.items():
            clean_key = k.strip().replace("\n", "").replace("\t", "").replace(" ", "")
            new[clean_key.lower()] = normalize_json_keys(v)
        return new
    if isinstance(obj, list):
        return [normalize_json_keys(x) for x in obj]
    return obj


# ==================================================
# Helpers de texto / Excel (Prompt 2 local)
# ==================================================
def normalize_text(s: str) -> str:
    if not isinstance(s, str):
        s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.replace("\xa0", " ").replace("\u200b", "")
    s = re.sub(r"\s+", " ", s)
    return s.strip().lower()


def get_materiality_table_from_excel(industry: str):
    """
    Lee la hoja 'ROI Final' de meteriality-map.xlsx y devuelve
    solo las filas donde Sector == industry (normalizado).
    NO usa pandas, solo openpyxl → evitamos numpy al 100%.
    """
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel no encontrado en: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)
    if "ROI Final" not in wb.sheetnames:
        raise ValueError("La hoja 'ROI Final' no existe en el Excel.")

    ws = wb["ROI Final"]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = next(rows_iter)
    except StopIteration:
        return {"materiality_table": [], "exhausted": True, "count": 0}

    # Mapear headers normalizados → índice de columna
    header_index = {}
    for idx, h in enumerate(headers):
        if h is None:
            continue
        header_index[normalize_text(str(h))] = idx

    def col_idx(header_name: str) -> Optional[int]:
        return header_index.get(normalize_text(header_name))

    # Índices esperados según tu Excel
    idx_sector = col_idx("sector")
    idx_temas = col_idx("temas")
    idx_mat_fin = col_idx("materialidad financiera")
    idx_val_mat_fin = col_idx("valor materialidad financiera")
    idx_riesgos = col_idx("riesgos")
    idx_oportunidades = col_idx("oportunidades")
    idx_acc_inicial = col_idx("accion inicial")  # "Acción inicial"
    idx_acc_moderada = col_idx("accion moderada")
    idx_acc_estructural = col_idx("accion estructural")

    missing = [
        name
        for name, idx in [
            ("Sector", idx_sector),
            ("Temas", idx_temas),
            ("Materialidad Financiera", idx_mat_fin),
            ("Valor materialidad financiera", idx_val_mat_fin),
            ("Riesgos", idx_riesgos),
            ("Oportunidades", idx_oportunidades),
            ("Acción inicial", idx_acc_inicial),
            ("Acción moderada", idx_acc_moderada),
            ("Acción estructural", idx_acc_estructural),
        ]
        if idx is None
    ]

    if missing:
        print(f"⚠ Columnas no encontradas en el Excel (normalizadas): {missing}")

    target = normalize_text(industry)
    result_rows = []

    for row in rows_iter:
        if idx_sector is None or idx_sector >= len(row):
            continue

        sector_val = row[idx_sector]
        if sector_val is None:
            continue

        if normalize_text(str(sector_val)) != target:
            continue

        def safe_get(idx_col):
            if idx_col is None or idx_col >= len(row):
                return None
            return row[idx_col]

        item = {
            # Claves igual que lo que venías usando en el JSON del assistant:
            "sector": safe_get(idx_sector),
            "temas": safe_get(idx_temas),
            "materialidadfinanciera": safe_get(idx_mat_fin),
            "valormaterialidadfinanciera": safe_get(idx_val_mat_fin),
            "riesgos": safe_get(idx_riesgos),
            "oportunidades": safe_get(idx_oportunidades),
            "accióninicial": safe_get(idx_acc_inicial),
            "acciónmoderada": safe_get(idx_acc_moderada),
            "acciónestructural": safe_get(idx_acc_estructural),
        }
        result_rows.append(item)

    print(f"✅ Filas encontradas en Excel para '{industry}': {len(result_rows)}")

    return {
        "materiality_table": result_rows,
        "exhausted": len(result_rows) == 0,
        "count": len(result_rows),
    }


async def run_prompt_assistant(message_text: str, use_tools: bool = False):
    # Crear thread
    thread = client.beta.threads.create()

    # Mandar mensaje del usuario
    client.beta.threads.messages.create(
        thread_id=thread.id,
        role="user",
        content=message_text,
    )

    # Ejecutar run con el assistant
    run_kwargs = {
        "thread_id": thread.id,
        "assistant_id": ASSISTANT_ID,
    }

    # 👇 Si NO queremos tools (Code Interpreter, etc.), las vaciamos
    if not use_tools:
        run_kwargs["tools"] = []

    run = client.beta.threads.runs.create(**run_kwargs)

    # Esperar a que termine (incluimos otros estados terminales por las dudas)
    while True:
        r = client.beta.threads.runs.retrieve(
            thread_id=thread.id,
            run_id=run.id,
        )
        if r.status in ("completed", "failed", "cancelled", "expired"):
            break
        await asyncio.sleep(1)

    # Obtener los mensajes finales del assistant
    messages = client.beta.threads.messages.list(thread_id=thread.id)

    print("Mensajes devueltos por el assistant (debug):")
    for msg in messages.data:
        print("ROLE:", msg.role)
        for c in msg.content:
            print("   - type:", getattr(c, "type", None))

    # Juntamos todos los bloques de texto de mensajes del assistant
    text_parts: list[str] = []

    # Recorremos de más antiguo a más nuevo
    for msg in reversed(messages.data):
        if msg.role != "assistant":
            continue

        for c in msg.content:
            # Muchos tipos (text, output_text, etc.) tienen .text
            text_obj = getattr(c, "text", None)
            if text_obj is not None:
                value = getattr(text_obj, "value", "")
                if value:
                    text_parts.append(value)

    if text_parts:
        full_text = "\n".join(text_parts)
        print("🔎 Texto combinado devuelto por assistant:", repr(full_text[:800]))
        return full_text

    print("⚠️ No se encontró texto en los mensajes del assistant.")
    return None


async def run_prompt_assistant_prompt10(message_text: str):
    """
    Versión especial de run_prompt_assistant para Prompt 10:
    - Fuerza el uso de Code Interpreter con tool_choice="required"
    - NO pisa la lista de tools del assistant (para no desactivar code_interpreter)
    """
    # Crear thread
    thread = client.beta.threads.create()

    # Mandar mensaje del usuario
    client.beta.threads.messages.create(
        thread_id=thread.id,
        role="user",
        content=message_text,
    )

    # Ejecutar run con el assistant, forzando tools
    run_kwargs = {
        "thread_id": thread.id,
        "assistant_id": ASSISTANT_ID,
        "tool_choice": "required",  # 👈 obliga a usar al menos un tool (code_interpreter)
    }

    run = client.beta.threads.runs.create(**run_kwargs)

    # Esperar a que termine (incluimos otros estados terminales por las dudas)
    while True:
        r = client.beta.threads.runs.retrieve(
            thread_id=thread.id,
            run_id=run.id,
        )
        if r.status in ("completed", "failed", "cancelled", "expired"):
            break
        await asyncio.sleep(1)

    # Obtener los mensajes finales del assistant
    messages = client.beta.threads.messages.list(thread_id=thread.id)

    print("Mensajes devueltos por el assistant (Prompt 10, debug):")
    for msg in messages.data:
        print("ROLE:", msg.role)
        for c in msg.content:
            print("   - type:", getattr(c, "type", None))

    # Juntamos todos los bloques de texto de mensajes del assistant
    text_parts: list[str] = []

    # Recorremos de más antiguo a más nuevo
    for msg in reversed(messages.data):
        if msg.role != "assistant":
            continue

        for c in msg.content:
            text_obj = getattr(c, "text", None)
            if text_obj is not None:
                value = getattr(text_obj, "value", "")
                if value:
                    text_parts.append(value)

    if text_parts:
        full_text = "\n".join(text_parts)
        print("🔎 Texto combinado devuelto por assistant (Prompt 10):", repr(full_text[:800]))
        return full_text

    print("⚠️ No se encontró texto en los mensajes del assistant para Prompt 10.")
    return None


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_SASB_PATH = os.path.join(BASE_DIR, "data", "lista_sasb.csv")

def load_sasb_rows_by_industry(industria_sasb: str):
    rows = []
    with open(CSV_SASB_PATH, mode="r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            if row["INDUSTRIA"].strip() == industria_sasb.strip():
                rows.append({
                    "industria": row["INDUSTRIA"],
                    "tema": row["TEMA"],
                    "parametro_contabilidad": row["PARÁMETRO DE CONTABILIDAD"],
                    "categoria": row["CATEGORÍA"],
                    "unidad_medida": row["UNIDAD DE MEDIDA"],
                    "codigo": row["CÓDIGO"],
                })
    return rows

def build_materiality_with_flag(p4, p5):
    """
    A partir de:
      - p4: JSON completo del Prompt 4 (materiality_table con TODAS las filas)
      - p5: JSON del Prompt 5 (materiality_table con SOLO los 10 temas materiales)
    construye una materiality_table completa con una columna 'tema_material'
    que marca cuáles son los 10 temas priorizados.
    """
    full_rows = (p4 or {}).get("materiality_table") or []
    top_rows = (p5 or {}).get("materiality_table") or []

    def _safe(s):
        if s is None:
            return ""
        return str(s)

    def _key(row):
        # Usamos combinación de sector + tema para identificar filas
        sector = _safe(row.get("sector", row.get("Sector")))
        tema = _safe(row.get("tema", row.get("Tema", row.get("temas", ""))))
        return normalize_text(sector) + "|||" + normalize_text(tema)

    top_keys = { _key(r) for r in top_rows }

    result = []
    for row in full_rows:
        new_row = dict(row)  # copia
        if _key(row) in top_keys:
            new_row["tema_material"] = "Tema Material"
        else:
            # Marcamos explícitamente como no material / NA
            new_row.setdefault("tema_material", "NA")
        result.append(new_row)

    print(f"✅ Filas totales en tabla completa: {len(result)}")
    print(f"✅ Filas marcadas como 'Tema Material': {len(top_keys)} (según Prompt 5)")
    return result



# ==================================================
# FUNCIÓN PRINCIPAL
# ==================================================
async def run_esg_analysis_prompt1_5(
    organization_name: str,
    country: str,
    website: str,
    industry: str,
    document: Optional[str] = None,
):
    print("\n🚀 Ejecutando Prompt 1 → Prompt 6…\n")

    responses = []
    failed_prompts: list[str] = []

    # ========= PROMPT 1 (via Assistant) =========
    print("\n📌 Ejecutando Prompt 1")
    p1_raw = prompt_1.format(
        organization_name=organization_name,
        country=country,
        website=website,
        industry=industry,
        document=document or "",
    )

    p1_text = await run_prompt_assistant(p1_raw)
    p1 = try_fix_json(p1_text) if p1_text else None

    responses.append({
        "name": prompt_1.name,
        "response_content": p1
    })

    if p1 is None:
        failed_prompts.append(prompt_1.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    # ========= PROMPT 2 (EXCEL local) =========
    print("\n📌 Ejecutando Prompt 2")
    excel_result = get_materiality_table_from_excel(industry)
    materiality_table = excel_result["materiality_table"]

    responses.append({
        "name": prompt_2.name,
        "response_content": {
            "materiality_table": materiality_table,
            "exhausted": len(materiality_table) == 0,
        },
    })

    if not materiality_table:
        failed_prompts.append(prompt_2.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    # ==================================================
    # A partir de acá vamos encadenando tabla → tabla
    # ==================================================

    # ========= PROMPT 3 (LLM con tabla de Prompt 2) =========
    print("\n📌 Ejecutando Prompt 3")
    p3_raw = (
        prompt_3.format()
        + "\n\nMATERIALITY_TABLE (JSON desde Prompt 2):\n"
        + "<MATERIALITY_TABLE_JSON>\n"
        + json.dumps({"materiality_table": materiality_table}, ensure_ascii=False)
        + "\n</MATERIALITY_TABLE_JSON>\n"
        + "\nRecuerda: usa EXACTAMENTE estas filas (mismo orden y cantidad)."
    )

    p3_text = await run_prompt_assistant(p3_raw)
    p3 = try_fix_json(p3_text) if p3_text else None

    responses.append({"name": prompt_3.name, "response_content": p3})

    if p3 is None:
        failed_prompts.append(prompt_3.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }


    # ========= PROMPT 4 (usa salida de Prompt 3) =========
    print("\n📌 Ejecutando Prompt 4")
    p3_json = json.dumps(p3, ensure_ascii=False)
    p4_raw = (
        prompt_4.format()
        + "\n\nMATERIALITY_TABLE (JSON desde Prompt 3):\n"
        + "<MATERIALITY_TABLE_JSON>\n"
        + p3_json
        + "\n</MATERIALITY_TABLE_JSON>\n"
        + "\nRecuerda: usa EXACTAMENTE estas filas (mismo orden y cantidad)."
    )

    p4_text = await run_prompt_assistant(p4_raw)
    p4 = try_fix_json(p4_text) if p4_text else None

    responses.append({"name": prompt_4.name, "response_content": p4})

    if p4 is None:
        failed_prompts.append(prompt_4.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    # ========= PROMPT 5 (usa salida de Prompt 4) =========
    print("\n📌 Ejecutando Prompt 5")
    p4_json = json.dumps(p4, ensure_ascii=False)
    p5_raw = (
        prompt_5.format()
        + "\n\nMATERIALITY_TABLE (JSON desde Prompt 4):\n"
        + "<MATERIALITY_TABLE_JSON>\n"
        + p4_json
        + "\n</MATERIALITY_TABLE_JSON>\n"
        + "\nRecuerda: ordena por materialidad_esg y selecciona SOLO 10 temas materiales."
    )

    p5_text = await run_prompt_assistant(p5_raw)
    p5 = try_fix_json(p5_text) if p5_text else None

    responses.append({"name": prompt_5.name, "response_content": p5})

    if p5 is None:
        failed_prompts.append(prompt_5.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

   # ========= PROMPT 6 (Code Interpreter + Excel ODS) =========
    print("\n📌 Ejecutando Prompt 6")

    p5_json = json.dumps(p5, ensure_ascii=False)

    try:
        # ⚠ En vez de usar prompt_6.format(...), tomamos el template crudo
        if hasattr(prompt_6, "template"):
            raw_template = prompt_6.template
        else:
            raw_template = str(prompt_6)

        # 👇 Reemplazo directo del placeholder, sin tocar el resto de llaves del JSON
        p6_raw = raw_template.replace("{materiality_table_json}", p5_json)
    except Exception as e:
        print("Error al preparar el prompt 6:", e)
        failed_prompts.append(prompt_6.name)
        responses.append({
            "name": prompt_6.name,
            "response_content": None,
        })
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    p6_text = await run_prompt_assistant(p6_raw, use_tools=True)

    p6 = try_fix_json_prompt6(p6_text) if p6_text else None

    responses.append({
        "name": prompt_6.name,
        "response_content": p6,
    })

    if p6 is None:
        failed_prompts.append(prompt_6.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }


    # ==================================================
    # PROMPT 8 (LLM + Code Interpreter → mapeo S&P → SASB)
    # ==================================================
    print("\n📌 Ejecutando Prompt 8")

    p8_raw = prompt_8.template.replace("{industry}", industry)

    # Igual que Prompt 6/10: forzamos tools
    p8_text = await run_prompt_assistant(p8_raw, use_tools=True)
    print("🔎 Texto combinado devuelto por assistant (Prompt 8):", repr(p8_text[:400]) if p8_text else None)

    p8_json = try_fix_json(p8_text) if p8_text else None

    responses.append({
        "name": prompt_8.name,
        "response_content": p8_json
    })

    # 🔐 Validación
    if not p8_json or "mapeo_sasb" not in p8_json:
        print("❌ Prompt 8 devolvió JSON inválido o sin clave 'mapeo_sasb'")
        failed_prompts.append(prompt_8.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    mapeo_sasb = p8_json.get("mapeo_sasb") or []

    if not isinstance(mapeo_sasb, list) or len(mapeo_sasb) == 0:
        print("❌ Prompt 8 devolvió 'mapeo_sasb' vacío (sin coincidencias SASB).")
        failed_prompts.append(prompt_8.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    if not isinstance(mapeo_sasb[0], dict) or "industria_sasb" not in mapeo_sasb[0]:
        print("❌ Prompt 8 devolvió 'mapeo_sasb[0]' sin 'industria_sasb'.")
        failed_prompts.append(prompt_8.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    industria_sasb = mapeo_sasb[0]["industria_sasb"]
    print(f"✅ Industria SASB detectada por Prompt 8: {industria_sasb}")



    # ==================================================
    # PROMPT 9 (CSV local — tabla SASB)
    # ==================================================
    print("\n📌 Ejecutando Prompt 9")

    tabla_sasb = load_sasb_rows_by_industry(industria_sasb)

    print(f"✅ CSV devolvió {len(tabla_sasb)} filas SASB para '{industria_sasb}'")

    if len(tabla_sasb) == 0:
        raise RuntimeError(
            f"❌ No se encontraron filas SASB para '{industria_sasb}'. "
            "Revisá si existe en lista_sasb.csv."
        )

    responses.append({
        "name": "🔹 Prompt 9: Mapeo SASB (CSV local)",
        "response_content": {
            "tabla_sasb": tabla_sasb
        }
    })

        


    # ========= PROMPT 10 (Code Interpreter + Excel mapeo_regulatorio) =========
    print("\n📌 Ejecutando Prompt 10")
    p5_json = json.dumps(p5, ensure_ascii=False)

    # inicializamos p10 para evitar NameError pase lo que pase
    p10 = None

    try:
        # 1) obtener template crudo
        if hasattr(prompt_10, "template"):
            raw_template_10 = prompt_10.template
        else:
            raw_template_10 = str(prompt_10)

        # 2) reemplazar placeholders
        p10_raw = (
            raw_template_10
            .replace("{materiality_table_json}", p5_json)
            .replace("{country}", country)
        )


        # 3) ejecutar assistant (¡usa SOLO la función especial de Prompt 10!)
        p10_text = await run_prompt_assistant_prompt10(p10_raw)

        # 4) parsear JSON
        p10 = try_fix_json_prompt10(p10_text) if p10_text else None

    except Exception as e:
        print("Error al ejecutar Prompt 10:", e)
        # dejamos p10 = None y manejamos abajo

    # 5) guardar respuesta en el array principal
    responses.append({
        "name": prompt_10.name,
        "response_content": p10,
    })

    if p10 is None:
        failed_prompts.append(prompt_10.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }


    # ========= PROMPT 11 (LLM puro, sin Code Interpreter) =========
    print("\n📌 Ejecutando Prompt 11")

    # Empaquetamos contexto relevante para que Prompt 11 tenga todo a mano
    p11_payload = {
        "prompt_1": p1,
        "prompt_2": {
            "materiality_table": materiality_table,
        },
        "prompt_3": p3,
        "prompt_4": p4,
        "prompt_5": p5,
        "prompt_6": p6,
        "prompt_8": p8_json,
        "prompt_9": {
            "tabla_sasb": tabla_sasb,
        },
        "prompt_10": p10,
    }

    # Construimos el texto del prompt 11 usando el template + contexto
    p11_raw = prompt_11.format(
        previous_results_json=json.dumps(p11_payload, ensure_ascii=False)
    )

    # Igual que Prompt 1: LLM "puro", sin tools
    p11_text = await run_prompt_assistant(p11_raw)
    p11 = try_fix_json(p11_text) if p11_text else None

    responses.append({
        "name": prompt_11.name,
        "response_content": p11,
    })

    if p11 is None:
        failed_prompts.append(prompt_11.name)
        return {
            "status": "partial",
            "responses": responses,
            "failed_prompts": failed_prompts,
        }

    # ========= TODO OK =========
    return {
        "status": "complete",
        "responses": responses,
        "failed_prompts": failed_prompts,
    }

